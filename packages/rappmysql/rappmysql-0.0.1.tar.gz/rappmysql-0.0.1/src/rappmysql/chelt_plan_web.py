import decimal
import os.path
import traceback
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5 import QtCore
from PyQt5.QtCore import *
import numpy as np
import datetime as dt
from datetime import datetime, timedelta
from dateutil.relativedelta import *
import sys
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import time
from mysqlquerys import connect

np.set_printoptions(linewidth=250)
__version__ = 'V4'


class CheltPlanificate:
    def __init__(self, ini_file, data_base_name):
        try:
            self.dataBase = connect.DataBase(ini_file, data_base_name)
        except FileNotFoundError as err:
            iniFile, a = QFileDialog.getOpenFileName(None, 'Open data base configuration file', os.getcwd(), "data base config files (*.ini)")
            if os.path.exists(iniFile):
                self.dataBase = connect.DataBase(iniFile, data_base_name)
            # ctypes.windll.user32.MessageBoxW(0, "Your text", "Your title", 1)
        except Exception as err:
            print(traceback.format_exc())

    def get_all_sql_vals(self, tableHead):
        # print(sys._getframe().f_code.co_name, tableHead)
        all_chelt = []
        for table in self.dataBase.tables:
            self.dataBase.active_table = table
            check = all(item in list(self.dataBase.active_table.columnsProperties.keys()) for item in tableHead)
            if check:
                vals = self.dataBase.active_table.returnColumns(tableHead)
                for row in vals:
                    row = list(row)
                    row.insert(0, table)
                    all_chelt.append(row)

        newTableHead = ['table']
        for col in tableHead:
            newTableHead.append(col)

        return newTableHead, all_chelt

    def filter_dates_old(self, tableHead, table, selectedStartDate, selectedEndDate):
        print(sys._getframe().f_code.co_name, tableHead)

        tableHead.append('payDay')
        validFromIndx = tableHead.index('valid_from')
        validToIndx = tableHead.index('valid_to')
        freqIndx = tableHead.index('freq')
        payDayIndx = tableHead.index('pay_day')
        postPayIndx = tableHead.index('post_pay')
        autoExtIndx = tableHead.index('auto_ext')
        nameIndx = tableHead.index('name')

        payments4Interval = []
        for val in table:
            # print(val)
            validfrom, validTo, freq, payDatum, postPay, autoExt = val[validFromIndx], \
                                                              val[validToIndx], \
                                                              val[freqIndx], \
                                                              val[payDayIndx], \
                                                              val[postPayIndx], \
                                                              val[autoExtIndx]

            if autoExt is None or autoExt == 0:
                autoExt = False
            else:
                autoExt = True
            if postPay is None or postPay == 0:
                postPay = False
            else:
                postPay = True

            #daca data expirarii este mai mica decat data de start selectata continua
            if validTo:
                if validTo < selectedStartDate and not autoExt:
                    if not postPay:
                        continue
            if not freq:
                continue

            if postPay:
                if not validTo:
                    paymentDate = datetime(validfrom.year, validfrom.month, payDatum).date() + relativedelta(months=freq)
                else:
                    paymentDate = validTo
            else:
                paymentDate = validfrom

            try:
                payDay = datetime(paymentDate.year, paymentDate.month, payDatum).date()
                if payDay < paymentDate:
                    payDay = datetime(paymentDate.year, paymentDate.month, payDatum).date() + relativedelta(months=1)
            except ValueError:
                payDay = datetime(paymentDate.year, paymentDate.month+1, 1).date() - relativedelta(days=1)
            except TypeError:
                payDay = paymentDate
            except Exception:
                print('OOOO')
                print(traceback.format_exc())
                sys.exit()

            toBePayed = False
            # cat timp data de end selectata este mai mare decat data platii...
            while selectedEndDate >= payDay:
                if selectedStartDate <= payDay <= selectedEndDate:
                    if not validTo:
                        tup = [x for x in val]
                        tup.append(payDay)
                        payments4Interval.append(tup)
                        toBePayed = True
                    elif payDay <= validTo:
                        tup = [x for x in val]
                        tup.append(payDay)
                        payments4Interval.append(tup)
                        toBePayed = True
                    elif payDay >= validTo and autoExt:
                        tup = [x for x in val]
                        tup.append(payDay)
                        payments4Interval.append(tup)
                        toBePayed = True
                    elif payDay >= validTo and postPay and selectedStartDate <= payDay <= selectedEndDate:
                        tup = [x for x in val]
                        tup.append(payDay)
                        payments4Interval.append(tup)
                        toBePayed = True

                payDay = payDay + relativedelta(months=+freq)
                try:
                    payDay = datetime(payDay.year, payDay.month, payDatum).date()
                except ValueError:
                    payDay = datetime(payDay.year, payDay.month + 1, 1).date() - relativedelta(days=1)
                except TypeError:
                    payDay = payDay
                except Exception:
                    print('OOOO')
                    print(traceback.format_exc())
                    sys.exit()
                # print(payDay, type(payDay), freq, type(freq), payDay.month+freq)
            if not toBePayed:
                continue
        payments4Interval = np.atleast_2d(payments4Interval)
        return tableHead, payments4Interval

    def filter_dates(self, tableHead, table, selectedStartDate, selectedEndDate):
        print(sys._getframe().f_code.co_name, tableHead)
        # print(sys._getframe().f_code.co_name, selectedStartDate, selectedEndDate)

        def get_next_pay_datum(validfrom, payDatum, freq, autoExt, validTo):
            dates2pay = []
            try:
                if postPay:
                    date_of_payment = datetime(validfrom.year, validfrom.month, payDatum).date() + relativedelta(months=freq)
                else:
                    date_of_payment = datetime(validfrom.year, validfrom.month, payDatum).date()
            except ValueError:
                if postPay:
                    date_of_payment = datetime(validfrom.year, validfrom.month + 1, 1).date() + relativedelta(months=freq) - timedelta(days=1)
                else:
                    date_of_payment = datetime(validfrom.year, validfrom.month + 1, 1).date() - timedelta(days=1)

            # print(date_of_payment, selectedStartDate <= date_of_payment <= selectedEndDate)
            if selectedStartDate <= date_of_payment < selectedEndDate:
                dates2pay.append(date_of_payment)

            if autoExt:
                while date_of_payment < selectedEndDate:
                    date_of_payment = date_of_payment + relativedelta(months=freq)
                    if date_of_payment.month == 3 and date_of_payment.day != payDatum:
                        date_of_payment = datetime(date_of_payment.year, date_of_payment.month, payDatum).date()
                    # print(date_of_payment, selectedStartDate <= date_of_payment <= selectedEndDate)
                    if selectedStartDate <= date_of_payment < selectedEndDate:
                        dates2pay.append(date_of_payment)
            elif validTo:
                if not autoExt and validTo > selectedEndDate:
                    while date_of_payment < selectedEndDate:
                        date_of_payment = date_of_payment + relativedelta(months=freq)
                        if date_of_payment.month == 3 and date_of_payment.day != payDatum:
                            date_of_payment = datetime(date_of_payment.year, date_of_payment.month, payDatum).date()
                        # print(date_of_payment, selectedStartDate <= date_of_payment <= selectedEndDate)
                        if selectedStartDate <= date_of_payment < selectedEndDate:
                            dates2pay.append(date_of_payment)

            return dates2pay

        tableHead.append('payDay')
        validFromIndx = tableHead.index('valid_from')
        validToIndx = tableHead.index('valid_to')
        freqIndx = tableHead.index('freq')
        payDayIndx = tableHead.index('pay_day')
        postPayIndx = tableHead.index('post_pay')
        autoExtIndx = tableHead.index('auto_ext')
        nameIndx = tableHead.index('name')

        payments4Interval = []
        for val in table:
            # print(val)
            validfrom, validTo, freq, payDatum, postPay, autoExt = val[validFromIndx], \
                                                              val[validToIndx], \
                                                              val[freqIndx], \
                                                              val[payDayIndx], \
                                                              val[postPayIndx], \
                                                              val[autoExtIndx]
            if autoExt is None or autoExt == 0:
                autoExt = False
            else:
                autoExt = True
            if postPay is None or postPay == 0:
                postPay = False
            else:
                postPay = True

            if not payDatum:
                continue
            dates_of_payment = get_next_pay_datum(validfrom, payDatum, freq, autoExt, validTo)
            if dates_of_payment:
                for date in dates_of_payment:
                    # print(str(date))
                    tup = []
                    for v in val:
                        tup.append(v)
                    tup.append(date)
                    payments4Interval.append(tuple(tup))

        payments4Interval = np.atleast_2d(payments4Interval)

        return tableHead, payments4Interval

    def filter_conto(self, tableHead, table, currentConto):
        # print(sys._getframe().f_code.co_name, tableHead, currentConto)
        if table.shape[1] > 0:
            if currentConto == 'all':
                indxConto = np.where(table[:, tableHead.index('table')] != 'intercontotrans')
            else:
                indxConto = np.where(table[:, tableHead.index('myconto')] == currentConto)
            return tableHead, table[indxConto]
        else:
            return tableHead, np.empty((0, len(tableHead)))

    def split_expenses_income(self, tableHead, table):
        # print(sys._getframe().f_code.co_name)
        indxValue = tableHead.index('value')
        payments = []
        income = []
        for row in table:
            if row[indxValue] > 0:
                income.append(row)
            if row[indxValue] < 0:
                payments.append(row)
        payments = np.atleast_2d(payments)
        income = np.atleast_2d(income)

        return payments, income


class MyApp(QMainWindow):
    def __init__(self):
        super(MyApp, self).__init__()
        path2src, pyFileName = os.path.split(__file__)
        uiFileName = 'chelt_plan.ui'
        path2GUI = os.path.join(path2src, 'GUI', uiFileName)
        Ui_MainWindow, QtBaseClass = uic.loadUiType(path2GUI)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        title = '{}_{}'.format(pyFileName, __version__)
        self.setWindowTitle(title)

        ini_file = r"D:\Python\MySQL\database.ini"
        data_base_name = 'cheltuieli'
        self.cheltPlan = CheltPlanificate(ini_file, data_base_name)
        self.tableHead = ['name', 'value', 'myconto', 'freq', 'pay_day', 'valid_from', 'valid_to', 'auto_ext', 'post_pay']

        self.myAccountsTable = connect.Table(ini_file, 'myfolderstructure', 'banca')
        self.myContos = self.myAccountsTable.returnColumn('name')

        self.populateCBConto()
        self.populateCBMonths()
        self.populateDatesInterval()
        self.prepareTablePlan()

        self.ui.cbActiveConto.currentIndexChanged.connect(self.prepareTablePlan)
        self.ui.CBMonths.currentIndexChanged.connect(self.populateDatesInterval)
        self.ui.DEFrom.dateTimeChanged.connect(self.prepareTablePlan)
        self.ui.DEBis.dateTimeChanged.connect(self.prepareTablePlan)
        self.ui.planTable.horizontalHeader().sectionClicked.connect(self.sortPlan)
        self.ui.PB_plotTablePie.clicked.connect(self.plotTablePie)
        self.ui.PB_plotNamePie.clicked.connect(self.plotNamePie)
        self.ui.PB_Plot.clicked.connect(self.plotGraf)
        self.ui.PB_export.clicked.connect(self.export)

    def export(self):
        expName, _ = QFileDialog.getSaveFileName(self, "Save file", "", "Excel Files (*.xlsx)")
        worksheets = [('Complete', datetime(datetime.now().year, 1, 1),datetime(datetime.now().year, 12, 31))]
        for mnth in range(1, 13):
            firstDayOfMonth = datetime(datetime.now().year, mnth, 1)
            if mnth != 12:
                lastDayOfMonth = datetime(datetime.now().year, mnth+1, 1) - timedelta(days=1)
            else:
                lastDayOfMonth = datetime(datetime.now().year + 1, 1, 1) - timedelta(days=1)

            tup = (firstDayOfMonth.strftime("%B"), firstDayOfMonth, lastDayOfMonth)
            worksheets.append(tup)

        wb = Workbook()
        ws = wb.active
        for mnth, firstDayOfMonth, lastDayOfMonth in worksheets:
            # print(mnth, firstDayOfMonth, lastDayOfMonth)
            if mnth == 'Complete':
                ws.title = mnth
            else:
                wb.create_sheet(mnth)
            ws = wb[mnth]
            self.ui.DEFrom.setDate(QDate(firstDayOfMonth))
            self.ui.DEBis.setDate(QDate(lastDayOfMonth))
            self.prepareTablePlan()

            planExpenseTable, planExpenseTableHead = self.readPlanExpenses()
            cheltData = np.insert(planExpenseTable, 0, planExpenseTableHead, 0)

            for i, row in enumerate(cheltData):
                for j, col in enumerate(row):
                    ws.cell(row=i + 1, column=j + 1).value = cheltData[i][j]

            firstRow = 1
            firstCol = get_column_letter(1)
            lastRow = len(cheltData)
            lastCol = get_column_letter(len(cheltData[0]))

            table_title = '{}_{}'.format('chelt', mnth )
            new_text = ('{}{}:{}{}'.format(firstCol, firstRow, lastCol, lastRow))
            tab = Table(displayName=table_title, ref=new_text)
            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)
            ws.cell(row=lastRow + 1, column=1).value = 'Total Number of Expenses'
            ws.cell(row=lastRow + 1, column=2).value = self.ui.LEtotalNoOfTransactions.text()
            ws.cell(row=lastRow + 2, column=1).value = 'Total Expenses'
            ws.cell(row=lastRow + 2, column=2).value = self.ui.LEtotalValue.text()
            #######income

            planIncomeTable, planIncomeTableHead = self.readPlanIncome()
            incomeData = np.insert(planIncomeTable, 0, planIncomeTableHead, 0)
            firstRow = lastRow + 5
            firstCol = get_column_letter(1)
            lastRow = firstRow + len(incomeData)
            lastCol = get_column_letter(len(incomeData[0]))

            for i, row in enumerate(incomeData):
                for j, col in enumerate(row):
                    ws.cell(row=i + firstRow, column=j + 1).value = incomeData[i][j]

            table_title = '{}_{}'.format('income', mnth )
            new_text1 = ('{}{}:{}{}'.format(firstCol, firstRow, lastCol, lastRow))
            tab = Table(displayName=table_title, ref=new_text1)
            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)
            ws.cell(row=lastRow + 1, column=1).value = 'Total Number of Incomes'
            ws.cell(row=lastRow + 1, column=2).value = self.ui.LEtotalNoOfIncome.text()
            ws.cell(row=lastRow + 2, column=1).value = 'Total Income'
            ws.cell(row=lastRow + 2, column=2).value = self.ui.LEtotalIncome.text()

        wb.save(expName)

    def populateCBMonths(self):
        self.ui.CBMonths.addItem('interval')
        months = [dt.date(2000, m, 1).strftime('%B') for m in range(1, 13)]
        for month in months:
            self.ui.CBMonths.addItem(month)

    def populateCBConto(self):
        print(sys._getframe().f_code.co_name)
        self.ui.cbActiveConto.addItem('all')
        self.ui.cbActiveConto.addItems(self.myContos)

    def populateDatesInterval(self):
        print(sys._getframe().f_code.co_name)
        startDate = QDate(datetime.now().year, datetime.now().month, datetime.now().day)
        if datetime.now().month != 12:
            mnth = datetime.now().month + 1
            lastDayOfMonth = datetime(datetime.now().year, mnth, 1) - timedelta(days=1)
        else:
            lastDayOfMonth = datetime(datetime.now().year + 1, 1, 1) - timedelta(days=1)

        if self.ui.CBMonths.currentText() != 'interval':
            mnth = datetime.strptime(self.ui.CBMonths.currentText(), "%B").month
            # print('****', mnth)
            # if mnth == 1:
            #     startDate = datetime(datetime.now().year - 1, 12, 30)
            # elif mnth == 2:
            #     startDate = datetime(datetime.now().year, mnth-1, 28)
            # else:
            #     startDate = datetime(datetime.now().year, mnth-1, 30)
            #
            # lastDayOfMonth = datetime(datetime.now().year, mnth, 29)

            startDate = datetime(datetime.now().year, mnth, 1)
            if mnth != 12:
                lastDayOfMonth = datetime(datetime.now().year, mnth+1, 1) - timedelta(days=1)
            else:
                lastDayOfMonth = datetime(datetime.now().year + 1, 1, 1) - timedelta(days=1)

            startDate = startDate - timedelta(days=2)
            lastDayOfMonth = lastDayOfMonth - timedelta(days=2)

            startDate = QDate(startDate)
            lastDayOfMonth = QDate(lastDayOfMonth)

        self.ui.DEFrom.setDate(startDate)
        self.ui.DEBis.setDate(lastDayOfMonth)

        self.ui.DEFrom.setCalendarPopup(True)
        self.ui.DEBis.setCalendarPopup(True)

    def prepareTablePlan(self):
        print(sys._getframe().f_code.co_name)
        currentConto = self.ui.cbActiveConto.currentText()

        selectedStartDate = self.ui.DEFrom.date()
        selectedEndDate = self.ui.DEBis.date()
        selectedStartDate = selectedStartDate.toPyDate()
        selectedEndDate = selectedEndDate.toPyDate()

        tableHead, table = self.cheltPlan.get_all_sql_vals(self.tableHead)

        tableHead, payments4Interval = self.cheltPlan.filter_dates(tableHead, table, selectedStartDate, selectedEndDate)

        tableHead, payments4Interval = self.cheltPlan.filter_conto(tableHead, payments4Interval, currentConto)

        payments4Interval, income = self.cheltPlan.split_expenses_income(tableHead, payments4Interval)

        if payments4Interval.shape == (1, 0):
            payments4Interval = np.empty((0, len(tableHead)))
        if income.shape == (1, 0):
            income = np.empty((0, len(tableHead)))

        self.populateExpensesPlan(tableHead, payments4Interval)
        self.populateTree(tableHead, payments4Interval)
        self.populateIncomePlan(tableHead, income)
        self.totals()

    def populateTree(self, tableHead, table):
        self.ui.TWmnthVSIrreg.clear()
        self.ui.TWmnthVSIrreg.setHeaderLabels(['freq', 'name', 'value'])
        monthly_level = QTreeWidgetItem(self.ui.TWmnthVSIrreg)
        monthly_level.setText(0, 'Monthly')
        irregular_level = QTreeWidgetItem(self.ui.TWmnthVSIrreg)
        irregular_level.setText(0, 'Irregular')
        monthlyIndx = np.where(table[:, tableHead.index('freq')] == 1)
        monthly = table[monthlyIndx]
        for mnth in monthly:
            mnth_item_level = QTreeWidgetItem(monthly_level)
            mnth_item_level.setText(1, mnth[tableHead.index('name')])
            mnth_item_level.setText(2, str(round(mnth[tableHead.index('value')])))

        totalMonthly = table[monthlyIndx,tableHead.index('value')][0]
        monthly_level.setText(1, 'Total')
        monthly_level.setText(2, str(round(sum(totalMonthly), 2)))

        irregIndx = np.where(table[:, tableHead.index('freq')] != 1)
        irregular = table[irregIndx]
        for irr in irregular:
            irr_item_level = QTreeWidgetItem(irregular_level)
            irr_item_level.setText(1, irr[tableHead.index('name')])
            irr_item_level.setText(2, str(round(irr[tableHead.index('value')], 2)))

        totalIrreg = table[irregIndx,tableHead.index('value')][0]
        irregular_level.setText(1, 'Total')
        irregular_level.setText(2, str(round(sum(totalIrreg), 2)))

    def populateExpensesPlan(self, tableHead, table):
        print(sys._getframe().f_code.co_name)
        self.ui.planTable.setColumnCount(len(tableHead))
        self.ui.planTable.setHorizontalHeaderLabels(tableHead)
        self.ui.planTable.setRowCount(table.shape[0])
        for col in range(table.shape[1]):
            for row in range(table.shape[0]):
                if isinstance(table[row, col], int) or isinstance(table[row, col], float):
                    item = QTableWidgetItem()
                    item.setData(QtCore.Qt.DisplayRole, table[row, col])
                elif isinstance(table[row, col], decimal.Decimal):
                    val = float(table[row, col])
                    item = QTableWidgetItem()
                    item.setData(QtCore.Qt.DisplayRole, val)
                else:
                    item = QTableWidgetItem(str(table[row, col]))
                self.ui.planTable.setItem(row, col, item)

        if table.shape[1] > 0:
            allValues = table[:, tableHead.index('value')]
            if None in allValues:
                allValues = allValues[allValues != np.array(None)]
            totalVal = round(sum(allValues.astype(float)), 2)
            self.ui.LEtotalNoOfTransactions.setText(str(len(table)))
            self.ui.LEtotalValue.setText(str(totalVal))

            indxMonthly = np.where(table[:,tableHead.index('freq')] == 1)[0]
            monthly = table[indxMonthly, tableHead.index('value')]
            if None in monthly:
                monthly = monthly[monthly != np.array(None)]
            totalMonthly = round(sum(monthly.astype(float)), 2)
            self.ui.LEnoOfMonthly.setText(str(monthly.shape[0]))
            self.ui.LEtotalMonthly.setText(str(totalMonthly))

            indxIrregular = np.where(table[:,tableHead.index('freq')] != 1)[0]
            irregular = table[indxIrregular, tableHead.index('value')]
            if None in irregular:
                irregular = irregular[irregular != np.array(None)]
            totalIrregular = round(sum(irregular.astype(float)), 2)
            self.ui.LEnoOfIrregular.setText(str(irregular.shape[0]))
            self.ui.LEirregular.setText(str(totalIrregular))

    def populateIncomePlan(self, tableHead, table):
        print(sys._getframe().f_code.co_name)
        self.ui.planTableIncome.setColumnCount(len(tableHead))
        self.ui.planTableIncome.setHorizontalHeaderLabels(tableHead)
        self.ui.planTableIncome.setRowCount(table.shape[0])
        for col in range(table.shape[1]):
            for row in range(table.shape[0]):
                if isinstance(table[row, col], int) or isinstance(table[row, col], float):
                    item = QTableWidgetItem()
                    item.setData(QtCore.Qt.DisplayRole, table[row, col])
                elif isinstance(table[row, col], decimal.Decimal):
                    val = float(table[row, col])
                    item = QTableWidgetItem()
                    item.setData(QtCore.Qt.DisplayRole, val)
                else:
                    item = QTableWidgetItem(str(table[row, col]))
                self.ui.planTableIncome.setItem(row, col, item)

        if table.shape[1] > 0:
            allValues = table[:, tableHead.index('value')]
            if None in allValues:
                allValues = allValues[allValues != np.array(None)]
            # for i in allValues:
            #     print(i, type(i))
            totalVal = sum(allValues.astype(float))
            self.ui.LEtotalNoOfIncome.setText(str(len(table)))
            self.ui.LEtotalIncome.setText(str(totalVal))

    def totals(self):
        if self.ui.LEtotalNoOfTransactions.text():
            expensesTrans = int(self.ui.LEtotalNoOfTransactions.text())
        else:
            expensesTrans = 0
        if self.ui.LEtotalNoOfIncome.text():
            incomeTrans = int(self.ui.LEtotalNoOfIncome.text())
        else:
            incomeTrans = 0

        if self.ui.LEtotalValue.text():
            expenses = float(self.ui.LEtotalValue.text())
        else:
            expenses = 0
        if self.ui.LEtotalIncome.text():
            income = float(self.ui.LEtotalIncome.text())
        else:
            income = 0

        trans = expensesTrans + incomeTrans
        total = round(expenses + income, 2)

        self.ui.LEtotalNo.setText(str(trans))
        self.ui.LEtotalVa.setText(str(total))

    def sortPlan(self, logical_index):
        print(sys._getframe().f_code.co_name)
        header = self.ui.planTable.horizontalHeader()
        order = Qt.DescendingOrder
        if not header.isSortIndicatorShown():
            header.setSortIndicatorShown(True)
        elif header.sortIndicatorSection() == logical_index:
            order = header.sortIndicatorOrder()
        header.setSortIndicator(logical_index, order)
        self.ui.planTable.sortItems(logical_index, order)

    def readPlanExpenses(self):
        rows = self.ui.planTable.rowCount()
        cols = self.ui.planTable.columnCount()
        planExpenseTable = np.empty((rows, cols), dtype=object)
        planExpenseTableHead = []
        for row in range(rows):
            for column in range(cols):
                cell = self.ui.planTable.item(row, column)
                planExpenseTable[row, column] = cell.text()
                colName = self.ui.planTable.horizontalHeaderItem(column).text()
                if colName not in planExpenseTableHead:
                    planExpenseTableHead.append(colName)

        return planExpenseTable, planExpenseTableHead

    def readPlanIncome(self):
        rows = self.ui.planTableIncome.rowCount()
        cols = self.ui.planTableIncome.columnCount()
        planIncomeTable = np.empty((rows, cols), dtype=object)
        planIncomeTableHead = []
        for row in range(rows):
            for column in range(cols):
                cell = self.ui.planTableIncome.item(row, column)
                planIncomeTable[row, column] = cell.text()
                colName = self.ui.planTableIncome.horizontalHeaderItem(column).text()
                if colName not in planIncomeTableHead:
                    planIncomeTableHead.append(colName)

        return planIncomeTable, planIncomeTableHead

    def plotTablePie(self):
        realExpenseTable, realExpenseTableHead = self.readPlanExpenses()
        allValues = realExpenseTable[:, realExpenseTableHead.index('value')].astype(float)
        if None in allValues:
            allValues = allValues[allValues != np.array(None)]
        totalVal = sum(allValues)

        colTableName = realExpenseTable[:, realExpenseTableHead.index('table')]
        labels = []
        sizes = []
        for table in np.unique(colTableName):
            indx = np.where(realExpenseTable[:, realExpenseTableHead.index('table')]==table)
            smallArray = realExpenseTable[indx]
            values = sum(smallArray[:, realExpenseTableHead.index('value')].astype(float))
            txt = '{} = {:.2f}'.format(table, values)
            labels.append(txt)
            size = (values/totalVal)*100
            sizes.append(size)

        fig1, ax1 = plt.subplots()
        ax1.pie(sizes, labels=labels, autopct='%1.2f%%', startangle=90)
        ax1.axis('equal')
        plt.legend(title='Total: {:.2f}'.format(totalVal))

        plt.show()

    def plotNamePie(self):
        realExpenseTable, realExpenseTableHead = self.readPlanExpenses()
        allValues = realExpenseTable[:, realExpenseTableHead.index('value')].astype(float)
        if None in allValues:
            allValues = allValues[allValues != np.array(None)]
        totalVal = sum(allValues)

        colTableName = realExpenseTable[:, realExpenseTableHead.index('name')]
        labels = []
        sizes = []
        for table in np.unique(colTableName):
            indx = np.where(realExpenseTable[:, realExpenseTableHead.index('name')]==table)
            smallArray = realExpenseTable[indx]
            values = sum(smallArray[:, realExpenseTableHead.index('value')].astype(float))
            txt = '{} = {:.2f}'.format(table, values)
            labels.append(txt)
            size = (values/totalVal)*100
            sizes.append(size)

        fig1, ax1 = plt.subplots()
        ax1.pie(sizes, labels=labels, autopct='%1.2f%%', startangle=90)
        ax1.axis('equal')
        plt.legend(title='Total: {:.2f}'.format(totalVal))

        plt.show()

    def plotGraf(self):
        realExpenseTable, realExpenseTableHead = self.readPlanExpenses()
        planIncomeTable, planIncomeTableHead = self.readPlanIncome()
        x_exp = []
        y_exp = []
        for date in np.unique(realExpenseTable[:, realExpenseTableHead.index('payDay')]):
            indx = np.where(realExpenseTable[:, realExpenseTableHead.index('payDay')] == date)
            arr = realExpenseTable[indx, realExpenseTableHead.index('value')].astype(float)
            x_exp.append(date)
            y_exp.append(abs(sum(arr[0])))

        x_inc = []
        y_inc = []
        for date in np.unique(planIncomeTable[:, planIncomeTableHead.index('payDay')]):
            indx = np.where(planIncomeTable[:, planIncomeTableHead.index('payDay')] == date)
            arr = planIncomeTable[indx, planIncomeTableHead.index('value')].astype(float)
            x_inc.append(date)
            y_inc.append(abs(sum(arr[0])))

        fig1, ax1 = plt.subplots()
        ax1.plot(x_exp, y_exp)
        ax1.plot(x_inc, y_inc)
        # plt.setp(plt.get_xticklabels(), rotation=30, ha="right")
        fig1.autofmt_xdate()
        plt.grid()
        plt.show()


def main():
    # app = QApplication(sys.argv)
    # window = MyApp()
    # window.show()
    # # sys.exit(app.exec_())
    # app.exec_()
    # ini_file = r"D:\Python\MySQL\database.ini"
    ini_file = r"D:\Python\MySQL\web_db.ini"
    # data_base_name = 'cheltuieli'
    data_base_name = 'heroku_6ed6d828b97b626'
    cheltPlan = CheltPlanificate(ini_file, data_base_name)
    tableHead = ['name', 'value', 'myconto', 'freq', 'pay_day', 'valid_from', 'valid_to', 'auto_ext', 'post_pay']
    newTableHead, all_chelt = cheltPlan.get_all_sql_vals(tableHead)

    for i in all_chelt:
        print(i)

if __name__ == '__main__':
    main()


